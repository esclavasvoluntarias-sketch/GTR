"""
Carga masiva por Excel: nómina de móviles/agentes y planificación semanal.
Formato esperado (fila de encabezado en la primera fila):

Nomina:       identificador | gps_device_id | base | tipo_servicio | agente_legajo | agente_nombre
Planificacion: identificador | fecha (YYYY-MM-DD) | hora_inicio | hora_fin | tipo_servicio | coeficiente
"""
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from datetime import datetime
import io

from app.core.database import get_db
from app.models.models import Movil, Agente, BaseDespacho, ServicioProyectado, TipoServicio

router = APIRouter(prefix="/api/carga", tags=["carga"])


def get_or_create_base(db: Session, codigo: str) -> BaseDespacho:
    base = db.query(BaseDespacho).filter(BaseDespacho.codigo == codigo).first()
    if not base:
        base = BaseDespacho(codigo=codigo, nombre=codigo)
        db.add(base)
        db.flush()
    return base


@router.post("/nomina")
async def cargar_nomina(archivo: UploadFile = File(...), db: Session = Depends(get_db)):
    if not archivo.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Subí un archivo .xlsx")

    contenido = await archivo.read()
    wb = load_workbook(io.BytesIO(contenido), data_only=True)
    ws = wb.active

    filas_procesadas = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        identificador, gps_device_id, base_codigo, tipo_servicio, legajo, nombre_agente = (row + (None,) * 6)[:6]

        base = get_or_create_base(db, str(base_codigo)) if base_codigo else None

        agente = None
        if legajo:
            agente = db.query(Agente).filter(Agente.legajo == str(legajo)).first()
            if not agente:
                agente = Agente(legajo=str(legajo), nombre=nombre_agente or str(legajo))
                db.add(agente)
                db.flush()

        movil = db.query(Movil).filter(Movil.identificador == str(identificador)).first()
        if not movil:
            movil = Movil(identificador=str(identificador))
            db.add(movil)

        movil.gps_device_id = str(gps_device_id) if gps_device_id else movil.gps_device_id
        movil.base_id = base.id if base else movil.base_id
        movil.agente_id = agente.id if agente else movil.agente_id
        try:
            movil.tipo_servicio = TipoServicio(tipo_servicio) if tipo_servicio else TipoServicio.OTRO
        except ValueError:
            movil.tipo_servicio = TipoServicio.OTRO

        filas_procesadas += 1

    db.commit()
    return {"filas_procesadas": filas_procesadas}


@router.post("/planificacion")
async def cargar_planificacion(archivo: UploadFile = File(...), db: Session = Depends(get_db)):
    if not archivo.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Subí un archivo .xlsx")

    contenido = await archivo.read()
    wb = load_workbook(io.BytesIO(contenido), data_only=True)
    ws = wb.active

    filas_procesadas = 0
    errores = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        identificador, fecha_str, hora_inicio, hora_fin, tipo_servicio, coeficiente = (row + (None,) * 6)[:6]

        movil = db.query(Movil).filter(Movil.identificador == str(identificador)).first()
        if not movil:
            errores.append(f"Fila {i}: móvil '{identificador}' no existe en la nómina")
            continue

        try:
            fecha = datetime.strptime(str(fecha_str), "%Y-%m-%d") if isinstance(fecha_str, str) else fecha_str
        except ValueError:
            errores.append(f"Fila {i}: fecha inválida '{fecha_str}'")
            continue

        servicio = ServicioProyectado(
            movil_id=movil.id,
            fecha=fecha,
            hora_inicio_turno=str(hora_inicio),
            hora_fin_turno=str(hora_fin),
            tipo_servicio=TipoServicio(tipo_servicio) if tipo_servicio in TipoServicio._value2member_map_ else TipoServicio.OTRO,
            coeficiente_operativo=float(coeficiente) if coeficiente else 1.0,
        )
        db.add(servicio)
        filas_procesadas += 1

    db.commit()
    return {"filas_procesadas": filas_procesadas, "errores": errores}
